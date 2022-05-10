#!/usr/bin/env python
# coding: utf-8

import os
import sys
import glob
import imghdr
import tempfile
import json
import traceback
import openpyxl
from logging import getLogger
from openpyxl.styles import Alignment
from PIL import Image
from .metadata import get_file_metadata, get_image_metadata


logger = getLogger(__name__)


def output_json(
    jsonpath: str,
    imgsdir: str,
    xlsxpath: str,
    recursive: bool,
    thumbssize: int,
    tags: list[str],
):
    """
    Generate a JSON file for CLI input arguments.

    Parameters
    ----------
    jsonpath : str
        Output JSON file path.
    imgspath : str
        Input directory that contain image files.
    xlsxpath: str
       Output Excel file name.
    thumbssize: int
       Thumbnails size.
    recursive: bool
        Recursively search for files.
    tags: list[str]
        Append exif tags.  The tag names may include group names, asusual in the format `<group>:<tag>`.
    """

    with open(jsonpath, mode="w", encoding="utf-8") as file:
        json.dump(
            {
                "inputdir": imgsdir,
                "output": xlsxpath,
                "recursive": recursive,
                "size": thumbssize,
                "tags": tags,
            },
            file,
            ensure_ascii=False,
            indent=2,
        )


def input_json(jsonpath: str):
    """
    Load CLI arguments from JSON file.

    Parameters
    ----------
    jsonpath : str
        Input JSON file path.
    """
    with open(jsonpath, mode="r", encoding="utf-8") as file:
        return json.load(file)


def _attach_image(ws, img: str, col: int, row: int):
    pilImage = Image.open(img)
    ws.row_dimensions[row].height = pilImage.height * 0.75
    wsImg = openpyxl.drawing.image.Image(img)
    cell_address = ws.cell(row=row, column=col).coordinate
    wsImg.anchor = cell_address
    ws.add_image(wsImg)
    return pilImage.width


def _add_tags(ws, tags, exif: dict, col: int, row: int):
    if len(tags) <= 0:
        return

    offset = 0
    for tag in tags:
        cell = ws.cell(row=row, column=col + offset)
        try:
            cell.value = str(exif.get(tag, ""))
            cell.alignment = Alignment(wrapText=True, vertical="top")
        except openpyxl.utils.exceptions.IllegalCharacterError as e:
            logger.warn(f"IllegalCharacterError: tag={tag}, value='{str(exif.get(tag))}'")
        offset += 1


def _retrieve_image_data(imgpath: str, size: int, outdir: str):

    if imghdr.what(imgpath) == None:
        return None, None

    try:
        pilImage = Image.open(imgpath)
        pilImage.thumbnail((size, size))
        path = os.path.join(outdir, os.path.basename(imgpath))
        pilImage.save(path)
    except Exception as e:
        logger.warn(f"failed to retrieve thumbsnail {str(e)}")
        return None, None

    metadata = {}

    get_file_metadata(imgpath, metadata)
    get_image_metadata(pilImage, metadata)

    return path, metadata


def run(
    imgspath: str,
    xlsxpath: str,
    thumbssize: int,
    tags: list[str],
    recursive: bool,
    callback=None,
):
    """
    Generate an Excel sheet with thumbnails from an image files.

    Parameters
    ----------
    imgspath : str
        Input directory that contain image files.
    xlsxpath: str
       Output Excel file name.
    thumbssize: int
       Thumbnails size.
    recursive: bool
        Recursively search for files.
    tags: list[str]
        Append exif tags.  The tag names may include group names, asusual in the format `<group>:<tag>`.
    callback: function
        The callback function that called when processed per file.
        ```
        def verbose_callback(filename, total, n):
            filename: processed file name.
            total: total files num.
            n: current file num.
        ```
    """
    imgspath = os.path.expanduser(imgspath)
    imgspath = os.path.expandvars(imgspath)

    xlsxpath = os.path.expanduser(xlsxpath)
    xlsxpath = os.path.expandvars(xlsxpath)

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "image list"

    ws.cell(1, 1).value = "No."
    ws.cell(1, 2).value = "Thumbnail"
    ws.cell(1, 3).value = "Filename"

    ws.freeze_panes = "C2"

    i = 4
    for tag in tags:
        ws.cell(1, i).value = tag
        i += 1

    files = sorted(glob.glob(os.path.join(imgspath, "**"), recursive=recursive))
    tmppath = tempfile.TemporaryDirectory()

    try:
        row = 2
        max_width = 0
        max_filename = 0

        filenum = len(files)

        for n, file in enumerate(files):
            if os.path.isdir(file):
                continue

            thumb, metadata = _retrieve_image_data(file, thumbssize, tmppath.name)

            if thumb != None:
                ws.cell(column=1, row=row).value = row - 1
                ws.cell(column=1, row=row).alignment = Alignment(vertical="top")
                width = _attach_image(ws, thumb, 2, row)
                if width > max_width:
                    max_width = width
                fn = os.path.basename(file)
                cell = ws.cell(column=3, row=row)
                cell.value = fn
                cell.alignment = Alignment(wrapText=True, vertical="top")

                if len(fn) > max_filename:
                    max_filename = len(fn)

                _add_tags(ws, tags, metadata, 4, row)

                row += 1

            if callback:
                callback(file, filenum, n + 1)

        ws.column_dimensions["B"].width = max_width * 0.13
        ws.column_dimensions["C"].width = (max_filename + 2) * 1.2

        for n in range(len(tags) + 1):
            colname = openpyxl.utils.get_column_letter(n + 3)
            max_length = 0
            for cell in ws[colname]:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            max_length = min(max_length, 100)
            ws.column_dimensions[colname].width = (max_length + 2) * 1.2

        wb.save(xlsxpath)
    except Exception as e:
        logger.error(traceback.format_exc())
    finally:
        tmppath.cleanup()
